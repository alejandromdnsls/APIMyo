# coding=utf-8
from openpyxl import Workbook
import glob
import sys

listOfFiles = glob.glob(sys.argv[1] + '*' + sys.argv[2] + '.csv')

wb = Workbook()
ws = wb.active
ws.title = "Data"

actual = 1

if len(listOfFiles) == 0:
    print "Error :" + sys.argv[2]

for nameFile in sorted(listOfFiles):
    #print nameFile
    nameCol = nameFile.split('\\')[-1].split(sys.argv[2])[0]
    file = open(nameFile, 'r')
    fila = 1
    for row in file:
        sum = actual + len(row.split('\n')[0].split(','))
        for col in range(len(row.split('\n')[0].split(','))):
            if fila != 1:
                ws.cell(row=fila, column=actual+col, value=float(row.split('\n')[0].split(',')[col]))
            else:
                if col == 0:
                    ws.cell(row=fila, column=actual+col, value=nameFile.split('\\')[-1].split('.csv')[0])
                else:
                    ws.cell(row=fila, column=actual+col, value=nameCol + row.split('\n')[0].split(',')[col])
        fila = fila + 1
    actual = sum
    file.close()

title = sys.argv[1].split('\\')[1]
title = title + sys.argv[2].split(title)[-1]
wb.save("MuestraNuevo\\" + title + ".xlsx")
