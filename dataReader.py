# coding=utf-8
from openpyxl import load_workbook

class dataReader:

    def __init__(self,route):
        self.wb = load_workbook(route)
        self.ws = self.wb.active

    def read(self):
        out = list()
        words = ['accelerometerx', 'accelerometery', 'accelerometerz', 'emgemg1', 'emgemg2', 'emgemg3', 'emgemg4', 'emgemg5', 'emgemg6', 'emgemg7', 'emgemg8']
        for col in self.ws.iter_cols(min_row=1):
            if col[0].value in words:
                aux = list()
                for cell in col:
                    if cell.value == None:
                            break
                    aux.append(cell.value)
                out.append(aux[1:])
        return out

if __name__ == '__main__':
    datos = dataReader('Datos//Muestra//a10.xlsx')
    print(len(datos.read()))
